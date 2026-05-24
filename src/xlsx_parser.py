from __future__ import annotations

import io
from typing import Any, Iterator

from openpyxl import load_workbook


def iter_xlsx_rows(content: bytes) -> Iterator[tuple[str, int, list[Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                yield sheet_name, idx, list(row)
    finally:
        wb.close()
