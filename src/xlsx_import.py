from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterator

from openpyxl import load_workbook

from src.snapshot_date import parse_snapshot_date
from src.currency_convert import normalize_v_record_currency
from src.product_format import format_import_light_fr, format_import_light_v, format_import_record

# F / R 表头映射（价格列名随汇率变化）
FR_FIELD_MAP = {
    "id": "item_id",
    "品牌": "brand",
    "型号": "model",
    "成色": "condition",
    "品相": "condition",
    "Condition": "condition",
    "condition": "condition",
    "颜色": "color",
    "图片路径": "image_path",
    "所有图片网址": "image_urls",
}

V_FIELD_MAP = {
    "图片": "image",
    "id": "item_id",
    "品牌": "brand",
    "商品名称": "product_name",
    "材质": "material",
    "颜色": "color",
    "尺寸": "size",
    "价格": "price",
    "货币": "currency",
    "卖家价格": "seller_price",
    "买家手续费": "buyer_fee",
    "点赞数": "likes",
    "上架时间": "listed_at",
    "售出时间": "sold_at",
    "成色": "condition",
    "品相": "condition",
    "Condition": "condition",
    "condition": "condition",
    "链接": "url",
}

FR_COLUMNS = [
    "item_id",
    "brand",
    "model",
    "condition",
    "price",
    "color",
    "material",
    "year",
    "other",
    "image_path",
    "image_urls",
    "old_data",
    "snapshot_date",
]

_FR_OLD_DATA_KEYS = ("item_id", "brand", "model", "condition", "color", "price")

_V_OLD_DATA_KEYS = (
    "item_id",
    "brand",
    "product_name",
    "material",
    "color",
    "condition",
    "size",
    "price",
    "currency",
    "seller_price",
    "buyer_fee",
    "likes",
    "listed_at",
    "sold_at",
    "url",
)


def build_old_data(record: dict[str, Any], table_kind: str) -> dict[str, Any]:
    """从 xlsx 解析行构建飞书原值快照（供 old_data 列与后续 LLM 回填）。"""
    keys = _FR_OLD_DATA_KEYS if table_kind in {"F", "R"} else _V_OLD_DATA_KEYS
    old_data: dict[str, Any] = {}
    for key in keys:
        val = record.get(key)
        if val is None:
            continue
        if isinstance(val, datetime):
            old_data[key] = val.isoformat()
        elif isinstance(val, (Decimal, int, float)):
            old_data[key] = str(val)
        else:
            old_data[key] = val
    return old_data


def prepare_import_record(record: dict[str, Any], table_kind: str) -> None:
    """导入入库：轻量规范化非型号列，型号及衍生列留空待 post-import 从 old_data 回填。"""
    if table_kind in {"F", "R"}:
        format_import_light_fr(record)
        record["model"] = None
        record["material"] = None
        record["year"] = None
        record["other"] = None
    elif table_kind == "V":
        format_import_light_v(record)
        record["product_name"] = None
        record["year"] = None
        record["other"] = None

def clear_image_fields(record: dict[str, Any], table_kind: str) -> None:
    """入库前清空图片相关列（后续可单独补 R2 URL）。"""
    if table_kind in {"F", "R"}:
        record["image_path"] = None
        record["image_urls"] = None
    elif table_kind == "V":
        record["image"] = None


V_COLUMNS = [
    "image",
    "item_id",
    "brand",
    "product_name",
    "material",
    "color",
    "year",
    "other",
    "size",
    "price",
    "currency",
    "seller_price",
    "buyer_fee",
    "likes",
    "listed_at",
    "sold_at",
    "condition",
    "url",
    "old_data",
    "snapshot_date",
]


def _normalize_header(cell: Any) -> str | None:
    if cell is None:
        return None
    text = str(cell).strip()
    if not text:
        return None
    upper = text.upper()
    if upper == "ID":
        return "item_id"
    if "价格" in text:
        return "price"
    return FR_FIELD_MAP.get(text) or V_FIELD_MAP.get(text) or V_FIELD_MAP.get(text.lower())


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _to_ts(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _row_to_fr(
    values: tuple[Any, ...],
    header_map: list[str | None],
    snapshot: date | None,
    *,
    skip_format: bool = False,
) -> dict[str, Any] | None:
    row: dict[str, Any] = {col: None for col in FR_COLUMNS}
    empty = True
    for idx, field in enumerate(header_map):
        if field is None or idx >= len(values):
            continue
        val = values[idx]
        if val is not None and str(val).strip() != "":
            empty = False
        if field == "item_id":
            row[field] = _to_str(val)
        elif field == "price":
            row[field] = _to_decimal(val)
        elif field == "condition":
            row[field] = _to_str(val)
        elif field == "color":
            row[field] = _to_str(val)
        else:
            row[field] = _to_str(val)
    if empty or not row.get("item_id"):
        return None
    row["snapshot_date"] = snapshot
    if not skip_format:
        format_import_record(row, table_kind="F")
    return row


def _row_to_v(
    values: tuple[Any, ...],
    header_map: list[str | None],
    snapshot: date | None,
    *,
    skip_format: bool = False,
) -> dict[str, Any] | None:
    row: dict[str, Any] = {col: None for col in V_COLUMNS}
    empty = True
    for idx, field in enumerate(header_map):
        if field is None or idx >= len(values):
            continue
        val = values[idx]
        if val is not None and str(val).strip() != "":
            empty = False
        if field == "item_id":
            row[field] = _to_str(val)
        elif field in {"price", "seller_price", "buyer_fee"}:
            row[field] = _to_decimal(val)
        elif field == "likes":
            row[field] = _to_int(val)
        elif field in {"listed_at", "sold_at"}:
            row[field] = _to_ts(val)
        elif field == "condition":
            row[field] = _to_str(val)
        elif field == "color":
            row[field] = _to_str(val)
        else:
            row[field] = _to_str(val)
    if empty or not row.get("item_id"):
        return None
    row["snapshot_date"] = snapshot
    normalize_v_record_currency(row)
    if not skip_format:
        format_import_record(row, table_kind="V")
    return row


def iter_sheet_rows(
    content: bytes,
    *,
    table_kind: str,
    snapshot_date: date | None,
    skip_format: bool = False,
) -> Iterator[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header_cells = next(rows_iter, None)
        if not header_cells:
            return
        header_map = [_normalize_header(c) for c in header_cells]

        convert = _row_to_fr if table_kind in {"F", "R"} else _row_to_v
        for values in rows_iter:
            if not values:
                continue
            record = convert(values, header_map, snapshot_date, skip_format=skip_format)
            if record:
                yield record
    finally:
        wb.close()
